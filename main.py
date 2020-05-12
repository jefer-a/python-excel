import cv2
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def toHex(rgb):
    color = ""
    color += str(hex(rgb[0])).replace('x', '0')[-2:]
    color += str(hex(rgb[1])).replace('x', '0')[-2:]
    color += str(hex(rgb[2])).replace('x', '0')[-2:]
    return color


def video2xlsx(cap, size=(0, 0), jump=0, outputPixelSize=(1, 1)):
    tabelNum = 1
    # while True:
    print('正在处理第' + str(tabelNum) + '帧...')
        # delay = jump
        # while delay != 0:
        #     ret, frame = cap.read()
        #     delay -= 1
    frame = cap
        # if size != (0, 0):
    
    w, h= frame.shape[:2]
    size=(int(h*0.5),int(w*0.5))
    frame = cv2.resize(frame, size)
    w, h= frame.shape[:2]
    wb = Workbook()
    ws = wb.active
    for i in range(h):
        ws.column_dimensions[get_column_letter(i+1)].width = outputPixelSize[0]
    for i in range(w):
        ws.row_dimensions[i+1].height = outputPixelSize[1]
    print("开始采集")
    print("h:",h,"w:",w)
    for i in range(h):
        for j in range(w):
            fill = PatternFill(fill_type="solid", fgColor=toHex(frame[j][i]))
            ws.cell(j + 1, i + 1).fill = fill
    print("结束采集")
    print("正在保存")
    try:
        wb.save('./outputs/' + str(tabelNum) + '.xlsx')
    except:
        print('保存失败，请检查文件是否已被打开')
    print("保存完成")
    if cv2.waitKey(40) & 0xFF == ord('q'):
        print('完成！')
        os.system('pause')
        # break
    tabelNum += 1


print('请输入视频文件绝对路径（或当前工作目录下的视频文件名）：')
test = 1
while test:
    capName = 'D:\\Github\\python-excel'
    if os.path.exists(capName):
        test = 0
    else:
        print('未找到文件，请重新输入：')

cap = cv2.imread('test.png')
sizeW = input('请输入缩小后宽度（默认不缩小）：')
sizeH = input('请输入缩小后高度（默认不缩小）：')
if sizeW == '' or sizeH == '':
    sizeW = '0'
    sizeH = '0'
jump = input('请设置每跳过多少帧采集一次（默认为0）：')
if jump == '':
    jump = '0'
outputkr = input('请设置输出单元格宽（默认为1）：')
if outputkr == '':
    outputkr = '1'
outputgc = input('请设置输出单元格高（默认为1）：')
if outputgc == '':
    outputgc = '1'

video2xlsx(cap, (int(sizeW), int(sizeH)), int(jump), (int(outputkr), int(outputgc)))
